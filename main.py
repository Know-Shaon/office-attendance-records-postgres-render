from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, session, jsonify
import psycopg2
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Protection

app = Flask(__name__)
app.secret_key = 'your_secret_key'

# Database connection
conn = psycopg2.connect(
    host="dpg-cq9f9k2ju9rs73b53rag-a.oregon-postgres.render.com",
    database="ford_office_attendance_records",
    user="ford_office_attendance_records_user",
    password="gLmdxUmTPA1JAgcDKOfBS9PIez0ugggt",
    port="5432"
)
cursor = conn.cursor()

# Ensure the Attendance Records directory exists
if not os.path.exists('Attendance Records'):
    os.makedirs('Attendance Records')

# Create or update tables
cursor.execute('''
CREATE TABLE IF NOT EXISTS teams (
    team_id TEXT PRIMARY KEY,
    team_name TEXT,
    member_password TEXT,
    manager_password TEXT
)
''')
cursor.execute('''
CREATE TABLE IF NOT EXISTS members (
    member_id SERIAL PRIMARY KEY,
    team_id TEXT,
    member_name TEXT,
    FOREIGN KEY(team_id) REFERENCES teams(team_id)
)
''')
cursor.execute('''
CREATE TABLE IF NOT EXISTS attendance (
    id SERIAL PRIMARY KEY,
    member_id INTEGER,
    date TEXT,
    status TEXT,
    is_manager BOOLEAN DEFAULT FALSE,
    FOREIGN KEY(member_id) REFERENCES members(member_id)
)
''')
conn.commit()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['POST'])
def login():
    team_id = request.form['team_id'].upper()
    password = request.form['password']
    cursor.execute('SELECT member_password, manager_password FROM teams WHERE team_id = %s', (team_id,))
    record = cursor.fetchone()
    if record:
        if record[0] == password:
            session['team_id'] = team_id
            session['is_manager'] = False
            return redirect(url_for('member_home'))
        elif record[1] == password:
            session['team_id'] = team_id
            session['is_manager'] = True
            return redirect(url_for('manager_home'))
    flash('Invalid credentials')
    return redirect(url_for('index'))

@app.route('/logout')
def logout():
    session.pop('team_id', None)
    session.pop('is_manager', None)
    return redirect(url_for('index'))

@app.route('/manager_home')
def manager_home():
    if 'team_id' in session and session.get('is_manager'):
        return render_template('manager_home.html', team_id=session['team_id'])
    else:
        flash('Unauthorized access')
        return redirect(url_for('index'))

@app.route('/member_home')
def member_home():
    if 'team_id' in session and not session.get('is_manager'):
        return render_template('member_home.html', team_id=session['team_id'])
    else:
        flash('Unauthorized access')
        return redirect(url_for('index'))

@app.route('/create_team', methods=['GET', 'POST'])
def create_team():
    if request.method == 'POST':
        team_name = request.form['team_name'].upper()
        member_password = request.form['member_password']
        manager_password = request.form['manager_password']
        cursor.execute('SELECT team_id FROM teams WHERE team_id = %s', (team_name,))
        existing_team = cursor.fetchone()
        if existing_team:
            flash('Team ID already exists')
        else:
            cursor.execute('''
            INSERT INTO teams (team_id, team_name, member_password, manager_password) VALUES (%s, %s, %s, %s)
            ''', (team_name, team_name, member_password, manager_password))
            conn.commit()
            flash('Team created successfully')
        return redirect(url_for('index'))
    return render_template('create_team.html')

@app.route('/add_member', methods=['GET', 'POST'])
def add_member():
    if 'team_id' in session and session.get('is_manager'):
        if request.method == 'POST':
            member_name = request.form['member_name']
            team_id = session['team_id']
            cursor.execute('SELECT member_id FROM members WHERE team_id = %s AND member_name = %s', (team_id, member_name))
            existing_member = cursor.fetchone()
            if existing_member:
                flash('Member already exists')
            else:
                cursor.execute('''
                INSERT INTO members (team_id, member_name) VALUES (%s, %s)
                ''', (team_id, member_name))
                conn.commit()
                flash('Member added successfully')
        return render_template('add_member.html', team_id=session['team_id'])
    else:
        flash('Unauthorized access')
        return redirect(url_for('index'))

@app.route('/remove_member', methods=['GET', 'POST'])
def remove_member():
    if 'team_id' in session and session.get('is_manager'):
        team_id = session['team_id']
        if request.method == 'POST':
            member_name = request.form['member_name']
            cursor.execute('SELECT member_id FROM members WHERE team_id = %s AND member_name = %s', (team_id, member_name))
            member_id = cursor.fetchone()
            if member_id:
                cursor.execute('DELETE FROM attendance WHERE member_id = %s', (member_id,))
                cursor.execute('DELETE FROM members WHERE member_id = %s', (member_id,))
                conn.commit()
                flash('Member and all their attendance records removed successfully')
            else:
                flash('Member not found')
        cursor.execute('SELECT member_name FROM members WHERE team_id = %s', (team_id,))
        members = cursor.fetchall()
        return render_template('remove_member.html', team_id=team_id, members=members)
    else:
        flash('Unauthorized access')
        return redirect(url_for('index'))

@app.route('/mark_attendance/<team_id>/<member_name>', methods=['GET', 'POST'])
def mark_attendance(team_id, member_name):
    team_id = team_id.upper()
    is_manager = session.get('is_manager', False)
    if request.method == 'POST':
        date = request.form['date']
        status = request.form['status']
        cursor.execute('SELECT member_id FROM members WHERE team_id = %s AND member_name = %s', (team_id, member_name))
        member_id = cursor.fetchone()
        if member_id:
            cursor.execute('SELECT * FROM attendance WHERE member_id = %s AND date = %s AND is_manager = %s', (member_id[0], date, is_manager))
            existing_record = cursor.fetchone()
            if existing_record:
                cursor.execute('''
                UPDATE attendance SET status = %s WHERE member_id = %s AND date = %s AND is_manager = %s
                ''', (status, member_id[0], date, is_manager))
            else:
                cursor.execute('''
                INSERT INTO attendance (member_id, date, status, is_manager) VALUES (%s, %s, %s, %s)
                ''', (member_id[0], date, status, is_manager))
            conn.commit()
            update_excel(team_id, is_manager)
            flash('Attendance marked successfully')
        else:
            flash('Member not found')
    return render_template('mark_attendance.html', team_id=team_id, member_name=member_name)

@app.route('/export_data')
def export_data():
    is_manager = session.get('is_manager', False)
    cursor.execute('SELECT * FROM attendance WHERE is_manager = %s', (is_manager,))
    records = cursor.fetchall()
    df = pd.DataFrame(records, columns=['ID', 'Member ID', 'Date', 'Status', 'Is Manager'])
    df.to_csv('attendance_records.csv', index=False)
    flash('Data exported to CSV successfully')
    return redirect(url_for('index'))

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory('Attendance Records', filename)

@app.route('/track_attendance/<team_id>', methods=['GET', 'POST'])
def track_attendance(team_id):
    team_id = team_id.upper()
    is_manager = session.get('is_manager', False)
    if request.method == 'POST':
        member_name = request.form['member_name']
        cursor.execute('''
        SELECT status, COUNT(*) FROM attendance
        JOIN members ON attendance.member_id = members.member_id
        WHERE members.team_id = %s AND members.member_name = %s AND attendance.is_manager = %s
        GROUP BY status
        ''', (team_id, member_name, is_manager))
        records = cursor.fetchall()
        stats = {status: count for status, count in records}
        return render_template('attendance_stats.html', team_id=team_id, member_name=member_name, stats=stats)
    cursor.execute('SELECT member_name FROM members WHERE team_id = %s', (team_id,))
    members = cursor.fetchall()
    return render_template('track_attendance.html', team_id=team_id, members=members)

@app.route('/download_monthly_report/<team_id>', methods=['GET', 'POST'])
def download_monthly_report(team_id):
    team_id = team_id.upper()
    is_manager = session.get('is_manager', False)
    if request.method == 'POST':
        month = request.form['month']
        year = request.form['year']
        cursor.execute('''
        SELECT t.team_name, m.member_name, a.date, a.status
        FROM attendance a
        JOIN members m ON a.member_id = m.member_id
        JOIN teams t ON m.team_id = t.team_id
        WHERE t.team_id = %s AND a.date LIKE %s AND a.is_manager = %s
        ''', (team_id, f'{year}-{month}%', is_manager))
        records = cursor.fetchall()
        df = pd.DataFrame(records, columns=['Team Name', 'Member Name', 'Date', 'Status'])

        filename = f'{team_id}_attendance_{year}_{month}.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Records"

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        for row in ws.iter_rows(min_row=1, min_col=1, max_row=1, max_col=len(df.columns)):
            for cell in row:
                cell.protection = Protection(locked=True)
                cell.font = Font(bold=True)

        ws.protection.sheet = True
        wb.save(os.path.join('Attendance Records', filename))

        return send_from_directory('Attendance Records', filename)

    return render_template('download_monthly_report.html', team_id=team_id)

@app.route('/admin_login')
def admin_login():
    return render_template('admin_login.html')

@app.route('/admin', methods=['POST'])
def admin():
    admin_password = request.form['admin_password']
    if admin_password == 'Admin@123':
        cursor.execute('SELECT team_id FROM teams')
        teams = cursor.fetchall()
        return render_template('admin.html', teams=teams)
    else:
        flash('Invalid admin password')
        return redirect(url_for('admin_login'))

@app.route('/remove_team', methods=['POST'])
def remove_team():
    team_id = request.form['team_id'].upper()
    cursor.execute('DELETE FROM attendance WHERE member_id IN (SELECT member_id FROM members WHERE team_id = %s)', (team_id,))
    cursor.execute('DELETE FROM members WHERE team_id = %s', (team_id,))
    cursor.execute('DELETE FROM teams WHERE team_id = %s', (team_id,))
    conn.commit()
    flash('Team and all its members and attendance records removed successfully')
    return redirect(url_for('admin'))

@app.route('/get_members/<team_id>')
def get_members(team_id):
    cursor.execute('SELECT member_name FROM members WHERE team_id = %s', (team_id,))
    members = cursor.fetchall()
    member_names = [member[0] for member in members]
    return jsonify({'members': member_names})

def update_excel(team_id, is_manager):
    cursor.execute('''
    SELECT t.team_name, m.member_name, a.date, a.status
    FROM attendance a
    JOIN members m ON a.member_id = m.member_id
    JOIN teams t ON m.team_id = t.team_id
    WHERE t.team_id = %s AND a.is_manager = %s
    ''', (team_id, is_manager))
    records = cursor.fetchall()
    df = pd.DataFrame(records, columns=['Team Name', 'Member Name', 'Date', 'Status'])

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Records"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.iter_rows(min_row=1, min_col=1, max_row=1, max_col=len(df.columns)):
        for cell in row:
            cell.protection = Protection(locked=True)
            cell.font = Font(bold=True)

    ws.protection.sheet = True
    filename = f'{team_id}_attendance_records.xlsx'
    wb.save(os.path.join('Attendance Records', filename))

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
